import csv
import math
from _datetime import datetime
import re
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from jinja2 import Environment, FileSystemLoader
import pdfkit
from openpyxl.reader.excel import load_workbook


currencyToRub = {
    "AZN": 35.68, "BYR": 23.91, "EUR": 59.90, "GEL": 21.74, "KGS": 0.76,
    "KZT": 0.13, "RUR": 1, "UAH": 1.64, "USD": 60.66, "UZS": 0.0055}


class Vacancy:
    def __init__(self, args):
        self.name = args[0]
        self.salary_from = float(args[1])
        self.salary_to = float(args[2])
        self.salary_currency = args[3]
        self.area_name = args[4]
        self.published_at = args[5]


class DataSet:
    def __init__(self, file_name):
        self.file_name = file_name
        self.vacancies_objects = list()

    @staticmethod
    def dataset(fileName):
        data = DataSet.csv_reader(fileName)
        dict = DataSet.csv_filter(data[0], data[1])
        dataset = DataSet(fileName)
        for item in dict:
            vacancy = Vacancy([f"{item['name']}", f"{item['salary_from']}", f"{item['salary_to']}",
                               f"{item['salary_currency']}", f"{item['area_name']}", f"{item['published_at']}"])
            vacancy.published_at = datetime.strptime(vacancy.published_at, "%Y-%m-%dT%H:%M:%S%z").year
            dataset.vacancies_objects.append(vacancy)
        return dataset

    @staticmethod
    def remove_tags(args):
        for i in range(len(args)):
            args[i] = " ".join(re.sub(r"\<[^>]*\>", "", args[i]).split())
        return args

    @staticmethod
    def csv_reader(filename):
        with open(filename, "r", encoding="utf-8-sig", newline="") as file:
            data = [x for x in csv.reader(file)]
        columns = data[0]
        rows = [x for x in data[1:] if len(x) == len(columns) and not x.__contains__("")]
        return columns, rows

    @staticmethod
    def csv_filter(columns, rows):
        dic_list = list()
        for row in rows:
            dic_result = dict()
            for i in range(len(row)):
                items = DataSet.remove_tags(row[i].split('\n'))
                dic_result[columns[i]] = items[0] if len(items) == 1 else "; ".join(items)
            dic_list.append(dic_result)
        return dic_list

class InputConnect:
    def __init__(self):
        self.params = InputConnect.get_prms()

    @staticmethod
    def get_prms():
        file_name = input("Введите название файла: ")
        vacancy = input("Введите название профессии: ")
        return file_name, vacancy

    @staticmethod
    def print(self, data: DataSet):
        def correctVacancy(data: DataSet):
            data.vacancy_by_city = {x: round(y / len(data.vacancies_objects), 4) for x, y in
                                         data.vacancy_by_city.items()}
            return dict(sorted(data.vacancy_by_city.items(), key=lambda item: item[1], reverse=True))
        data.vacancies_count_by_year = InputConnect.vacanciesName(data, "None")
        data.salary_by_year = InputConnect.SalaryName(data, "None")
        data.vacancies_count_by_profession_name = InputConnect.vacanciesName(data, self.params[1])
        data.salary_by_profession_name = InputConnect.SalaryName(data, self.params[1])
        data.vacancy_by_city = InputConnect.VacancyCity(data)
        data.salary_by_city = InputConnect.SalaryCity(data)
        data.vacancy_by_city = correctVacancy(data)
        data.dict_lict = [data.salary_by_year, data.salary_by_profession_name, data.vacancies_count_by_year,
                          data.vacancies_count_by_profession_name, dict(list(data.salary_by_city.items())[:10]),
                          data.vacancy_by_city]
        print(f"Динамика уровня зарплат по годам: {data.salary_by_year}")
        print(f"Динамика количества вакансий по годам: {data.vacancies_count_by_year}")
        print(f"Динамика уровня зарплат по годам для выбранной профессии: {data.salary_by_profession_name}")
        print(f"Динамика количества вакансий по годам для выбранной профессии: {data.vacancies_count_by_profession_name}")
        print(f"Уровень зарплат по городам (в порядке убывания): {dict(list(data.salary_by_city.items())[:10])}")
        print(f"Доля вакансий по городам (в порядке убывания): {dict(list(data.vacancy_by_city.items())[:10])}")

    @staticmethod
    def vacanciesName(data: DataSet, name):
        count = dict()
        for vac in data.vacancies_objects:
            if vac.name.__contains__(name) or name == "None":
                InputConnect.set_name(count, vac.published_at)
        if len(count) == 0:
            return {2022: 0}
        return count

    @staticmethod
    def SalaryName(data: DataSet, name):
        salary = dict()
        for vac in data.vacancies_objects:
            if vac.name.__contains__(name) or name == "None":
                if not salary.__contains__(vac.published_at):
                    salary[vac.published_at] = InputConnect.currencyRub(vac)
                else:
                    salary[vac.published_at] += InputConnect.currencyRub(vac)
        if len(salary) == 0:
            return {2022: 0}
        for key in salary.keys():
            if name == "None":
                salary[key] = math.floor(salary[key] / data.vacancies_count_by_year[key])
            else:
                salary[key] = math.floor(salary[key] / data.vacancies_count_by_profession_name[key])
        return salary

    @staticmethod
    def VacancyCity(data: DataSet):
        vacancy = dict()
        for vac in data.vacancies_objects:
            InputConnect.set_name(vacancy, vac.area_name)
        return vacancy

    @staticmethod
    def SalaryCity(data: DataSet):
        salary = dict()
        for vac in data.vacancies_objects:
            if math.floor(data.vacancy_by_city[vac.area_name] / len(data.vacancies_objects) * 100) >= 1:
                if not salary.__contains__(vac.area_name):
                    salary[vac.area_name] = InputConnect.currencyRub(vac)
                else:
                    salary[vac.area_name] += InputConnect.currencyRub(vac)
        for key in salary:
            salary[key] = math.floor(salary[key] / data.vacancy_by_city[key])
        return dict(sorted(salary.items(), key=lambda item: item[1], reverse=True))

    @staticmethod
    def set_name(dict: dict, name):
        if not dict.__contains__(name):
            dict[name] = 1
        else:
            dict[name] += 1

    @staticmethod
    def currencyRub(vacancy):
        course = currencyToRub[vacancy.salary_currency]
        return int((vacancy.salary_from * course + vacancy.salary_to * course) / 2)


class Report:
    def __init__(self, dict_lict: list()):
        self.data = dict_lict

    def reportExcel(self, name):
        def as_text(value):
            if value is None:
                return ""
            return str(value)

        def set_max_length(worksheet):
            for column_cells in worksheet.columns:
                length = max(len(as_text(cell.value)) for cell in column_cells)
                worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

        def set_format_percent(worksheet):
            for i, column_cells in enumerate(worksheet.columns):
                if i == 4:
                    for cell in column_cells:
                        cell.number_format = FORMAT_PERCENTAGE_00

        def set_headers(headers, head_range):
            for i, cell in enumerate(head_range):
                cell.value = headers[i]
                cell.font = Font(size=11, b=True)

        def set_border_style(worksheet):
            for column_cells in worksheet.columns:
                for cell in column_cells:
                    bd = Side(style="thin", color="000000")
                    cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)

        wb = Workbook()
        sheet_1 = wb.worksheets[0]
        sheet_1.title = "Статистика по годам"
        sheet_2 = wb.create_sheet("Статистика по городам")
        headers = ["Год", "Средняя зарплата", f"Средняя зарплата - {name}",
                   "Количество вакансий", f"Количество вакансий - {name}"]
        set_headers(headers, sheet_1['A1':'E1'][0])
        for key in self.data[0].keys():
            sheet_1.append([key, self.data[0][key], self.data[1][key], self.data[2][key], self.data[3][key]])
        set_border_style(sheet_1)
        set_max_length(sheet_1)
        set_headers(["Город", "Уровень зарплат"], sheet_2['A1':'B1'][0])
        set_headers(["Город", "Доля вакансий"], sheet_2['D1':'E1'][0])
        sheet_2.column_dimensions['C'].width = 2
        city_keys = list(self.data[5].keys())
        for i, key in enumerate(self.data[4].keys()):
            sheet_2.append([key, self.data[4][key], None, city_keys[i], self.data[5][city_keys[i]]])
        for i, column_cells in enumerate(sheet_2.columns):
            for cell in column_cells:
                if i != 2:
                    bd = Side(style="thin", color="000000")
                    cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        set_format_percent(sheet_2)
        set_max_length(sheet_2)
        wb.save("report.xlsx")
        return

    def reportImage(self, name):
        def myfunc(item):
            if item.__contains__(' '):
                return item[:item.index(' ')] + '\n' + item[item.index(' ')+1:]
            elif item.__contains__('-'):
                return item[:item.index('-')] + '-\n' + item[item.index('-') + 1:]
            return item

        width = 0.3
        nums = np.arange(len(self.data[0].keys()))
        dx1 = nums - width / 2
        dx2 = nums + width / 2
        fig = plt.figure()
        ax = fig.add_subplot(221)
        ax.set_title("Уровень зарплат по годам")
        ax.bar(dx1, self.data[0].values(), width, label="средняя з/п")
        ax.bar(dx2, self.data[1].values(), width, label=f"з/п {name.lower()}")
        ax.set_xticks(nums, self.data[0].keys(), rotation="vertical")
        ax.legend(fontsize=8)
        ax.tick_params(axis="both", labelsize=8)
        ax.grid(True, axis='y')
        ax = fig.add_subplot(222)
        ax.set_title("Количество вакансии по годам")
        ax.bar(dx1, self.data[2].values(), width, label="Количество вакансии")
        ax.bar(dx2, self.data[3].values(), width, label=f"Количество вакансии\n{name.lower()}")
        ax.set_xticks(nums, self.data[0].keys(), rotation="vertical")
        ax.legend(fontsize=8)
        ax.tick_params(axis="both", labelsize=8)
        ax.grid(True, axis='y')
        ax = fig.add_subplot(223)
        ax.set_title("Уровень зарплат по городам")
        cities = list(map(myfunc, tuple(self.data[4].keys())))
        y_pos = np.arange(len(cities))
        ax.barh(y_pos, list(self.data[4].values()), align='center')
        ax.set_yticks(y_pos, labels=cities)
        ax.invert_yaxis()
        ax.grid(True, axis='x')
        ax = fig.add_subplot(224)
        ax.set_title("Доля вакансии по годам")
        labels = list(dict(list(self.data[5].items())[:10]).keys())
        labels.insert(0, "Другое")
        vals = list(dict(list(self.data[5].items())[:10]).values())
        vals.insert(0, 1 - sum(list(dict(list(self.data[5].items())[:10]).values())))
        ax.pie(vals, labels=labels, startangle=0, textprops={"fontsize": 6})
        plt.tight_layout()
        fig.set_size_inches(9.5, 7.5)
        plt.savefig("graph.png", dpi=120)
        return

    def make_pdf(self, name):
        self.reportExcel(name)
        self.reportImage(name)
        image_file = "graph.png"
        book = load_workbook("report.xlsx")
        sheet_1 = book.active
        sheet_2 = book['Статистика по городам']
        for row in range(2, sheet_2.max_row + 1):
            for col in range(4, 6):
                if type(sheet_2.cell(row, col).value).__name__ == "float":
                    sheet_2.cell(row, col).value = str(round(sheet_2.cell(row, col).value * 100, 2)) + '%'
        options = {'enable-local-file-access': None}
        en = Environment(loader=FileSystemLoader('.'))
        template = en.get_template("template.html")
        template = template.render({'name': name, 'image_file': image_file, 'sheet_1': sheet_1, 'sheet_2': sheet_2})
        config = pdfkit.configuration(wkhtmltopdf=r'D:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(template, 'Report.pdf', configuration=config, options=options)

inputParam = InputConnect()
dataSet = DataSet.dataset(inputParam.params[0])
InputConnect.print(inputParam, dataSet)
report = Report(dataSet.dict_lict)
report.make_pdf(inputParam.params[1])